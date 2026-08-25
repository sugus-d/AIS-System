#!/usr/bin/env python3
"""生成 Table 1-6 Excel 文件，对齐 HTML 版本布局和数据。

用法:
    uv run python -m commands.export.excel
"""

from __future__ import annotations

import math

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.stats import f_oneway, pearsonr

from commands.export.config import TABLES_DIR
from reports.pages.data_tables.data import load_clinical_stats
from utils.logger import logger

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BFONT = Font(bold=True, size=11)
NFONT = Font(size=11)
HFILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Cobb 角严重度分级阈值
_COBB_MILD = 10
_COBB_MODERATE = 20
_COBB_SEVERE = 40
# ANOVA 最少分组数
_MIN_ANOVA_GROUPS = 2
# Pearson 相关最少样本数
_MIN_PEARSON_SAMPLES = 3
# p 值显示阈值：小于此值改用科学计数法
_P_VALUE_TINY = 0.0001


def sc(cell, font=NFONT, align=CENTER):
    cell.font = font
    cell.alignment = align
    cell.border = BORDER


def sh(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BFONT
        cell.alignment = CENTER
        cell.border = BORDER
        cell.fill = HFILL


def make_table1():
    clinical = load_clinical_stats()
    df = pd.read_csv(TABLES_DIR / "table1_raw.csv")
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 1"

    for i, h in enumerate(["Characteristic",
        "Total\n(n=122)", "Non-AIS", "AIS",
        "Male\n(n=32)", "Non-AIS", "AIS",
        "Female\n(n=90)", "Non-AIS", "AIS"], 1):
        ws.cell(row=1, column=i, value=h)
    sh(ws, 1, 1, 10)
    ws.row_dimensions[1].height = 30

    bmi_d = clinical.get("bmi_detail", {})
    ct_d = clinical.get("ct_detail", {})
    sev_d = clinical.get("sev_detail", {})

    def ms(gender: str, ais: str, col: str) -> str:
        sub = df[df["ais"] == ais] if gender == "Total" else df[(df["gender"] == gender) & (df["ais"] == ais)]
        vals = sub[col].dropna().values.astype(float)
        return f"{vals.mean():.1f}±{vals.std():.1f}" if len(vals) else "-"

    def combine_total(col: str) -> str:
        vals = df[col].dropna().values.astype(float)
        return f"{vals.mean():.1f}±{vals.std():.1f}" if len(vals) else "-"

    def combine_v(col: str) -> list[str]:
        total = combine_total(col)
        return [total, ms("Total", "Non-AIS", col), ms("Total", "AIS", col),
                f"{ms('Male', 'Non-AIS', col)} / {ms('Male', 'AIS', col)}",
                ms("Male", "Non-AIS", col), ms("Male", "AIS", col),
                f"{ms('Female', 'Non-AIS', col)} / {ms('Female', 'AIS', col)}",
                ms("Female", "Non-AIS", col), ms("Female", "AIS", col)]

    def cat_cells(detail: dict, label: str) -> list[int]:
        data = detail.get(label, {})
        def gv(g: str, a: str) -> int:
            return data.get(g, {}).get(a, 0)
        all_non = gv("Male","Non-AIS") + gv("Female","Non-AIS")
        all_ais = gv("Male","AIS") + gv("Female","AIS")
        return [all_non + all_ais, all_non, all_ais,
                gv("Male","Non-AIS") + gv("Male","AIS"), gv("Male","Non-AIS"), gv("Male","AIS"),
                gv("Female","Non-AIS") + gv("Female","AIS"), gv("Female","Non-AIS"), gv("Female","AIS")]

    rows = []
    rows.append(("Age (years)", ["-"] * 9))
    rows.append(("Height (cm)", combine_v("height_cm")))
    rows.append(("Weight (kg)", combine_v("weight_kg")))
    rows.append(("BMI (kg/m²)", combine_v("bmi")))
    for label in ["Underweight", "Normal weight", "Overweight", "Obesity"]:
        rows.append((f"  {label}", cat_cells(bmi_d, label)))
    rows.append(("Cobb angle (°)", combine_v("max_cobb")))
    rows.append(("Curve Type", [""]*9))
    for ct_key, ct_label in [("Thoracic", "  Thoracic curve"), ("Lumbar", "  Lumbar curve"), ("Double", "  Double Curves")]:
        rows.append((ct_label, cat_cells(ct_d, ct_key)))
    rows.append(("Severity", [""]*9))
    for s in ["Normal", "Mild", "Moderate", "Severe"]:
        rows.append((f"  {s}", cat_cells(sev_d, s)))

    for ri, (label, vals) in enumerate(rows, 2):
        c = ws.cell(row=ri, column=1, value=label)
        is_sub = label.startswith("  ")
        sc(c, font=BFONT if label in ("Curve Type", "Severity") else NFONT, align=LEFT)
        if is_sub:
            c.font = NFONT
        for j, v in enumerate(vals, 2):
            sc(ws.cell(row=ri, column=j, value=v if v is not None else ""))

    ws.column_dimensions["A"].width = 22
    for c in range(2, 11):
        ws.column_dimensions[get_column_letter(c)].width = 14
    wb.save(TABLES_DIR / "table1.xlsx")
    logger.info("Table 1 OK")


def make_table2():
    df = pd.read_csv(TABLES_DIR / "table2_raw.csv")
    cols = ["Sh.IB", "Sh.A", "Sca.IB", "Sca.A", "ASIS.A", "Trunk.L", "Sh.W", "Sh.AI", "Pe.AI"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 2"
    headers = ["Index", "Auto", "Manual", "Observer 1", "Observer 2",
               "MAE", "RMSE", "R²", "r", "ICCa", "ICCb"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    sh(ws, 1, 1, 11)
    for ri, col in enumerate(cols, 2):
        vals = df[col].dropna().values.astype(float)
        auto = f"{vals.mean():.2f}±{vals.std():.2f}" if len(vals) else ""
        sc(ws.cell(row=ri, column=1, value=col), align=LEFT)
        sc(ws.cell(row=ri, column=2, value=auto))
        for j in range(3, 12):
            sc(ws.cell(row=ri, column=j, value=""))
    ws.column_dimensions["A"].width = 14
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 14
    wb.save(TABLES_DIR / "table2.xlsx")
    logger.info("Table 2 OK")


def make_table3():
    df = pd.read_csv(TABLES_DIR / "table3_raw.csv")
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 3"
    headers = ["Index", "Whole cohort", "Normal", "Mild", "Moderate", "Severe", "P value"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    sh(ws, 1, 1, 7)

    df["severity"] = df["max_cobb"].apply(
        lambda x: "Normal" if x < _COBB_MILD else "Mild" if x < _COBB_MODERATE else "Moderate" if x < _COBB_SEVERE else "Severe")
    idx_map = {"ai": "Asymmetric Index", "curvature_index": "Curvature Index", "height_index": "Height Index", "nai": "Normal Angle Index", "ri": "Roughness Index"}
    sev_order = ["Whole cohort", "Normal", "Mild", "Moderate", "Severe"]

    ri = 2
    for idx_key, idx_label in idx_map.items():
        sc(ws.cell(row=ri, column=1, value=idx_label), align=LEFT)
        for j, sev_name in enumerate(sev_order, 2):
            sub = df if sev_name == "Whole cohort" else df[df["severity"] == sev_name]
            vals = sub[idx_key].dropna().values.astype(float)
            val_str = f"{vals.mean():.1f}±{vals.std():.1f}" if len(vals) else "-"
            sc(ws.cell(row=ri, column=j, value=val_str))
        groups = []
        for sev_name in ["Normal", "Mild", "Moderate", "Severe"]:
            vals = df[df["severity"] == sev_name][idx_key].dropna().values.astype(float)
            if len(vals) > 1:
                groups.append(vals)
        p_val = f_oneway(*groups)[1] if len(groups) >= _MIN_ANOVA_GROUPS else 1.0
        sc(ws.cell(row=ri, column=7, value=f"{p_val:.4f}" if p_val > _P_VALUE_TINY else f"{p_val:.2e}"))
        ri += 1

    ri += 1
    ws.column_dimensions["A"].width = 22
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 18
    wb.save(TABLES_DIR / "table3.xlsx")
    logger.info("Table 3 OK")


def make_table4():
    df = pd.read_csv(TABLES_DIR / "table4_raw.csv")
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 4"
    headers = ["Index", "Severity", "Correlation coefficient (r)", "p-value"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    sh(ws, 1, 1, 4)

    df["severity"] = df["max_cobb"].apply(
        lambda x: "Normal" if x < _COBB_MILD else "Mild" if x < _COBB_MODERATE else "Moderate" if x < _COBB_SEVERE else "Severe")
    idx_map = {"ai": "Asymmetric Index", "curvature_index": "Curvature Index", "height_index": "Height Index", "nai": "Normal Angle Index", "ri": "Roughness Index"}
    sev_sub = {k: (df if k == "All" else df[df["severity"] == k])
               for k in ["All", "Normal", "Mild", "Moderate", "Severe"]}
    sev_names = {"All": "Whole cohort", "Normal": "Normal cases", "Mild": "Mild cases",
                  "Moderate": "Moderate cases", "Severe": "Severe cases"}

    ri = 2
    for idx_key, idx_label in idx_map.items():
        first = True
        for sev_key in ["All", "Normal", "Mild", "Moderate", "Severe"]:
            sub = sev_sub[sev_key].dropna(subset=[idx_key])
            vals = sub[idx_key].values.astype(float)
            y_sub = sub["max_cobb"].values.astype(float)
            if first:
                sc(ws.cell(row=ri, column=1, value=idx_label), align=LEFT)
                first = False
            sc(ws.cell(row=ri, column=2, value=sev_names[sev_key]))
            if len(vals) >= _MIN_PEARSON_SAMPLES:
                r, p = pearsonr(vals, y_sub)
                sc(ws.cell(row=ri, column=3, value=round(r, 4)))
                sc(ws.cell(row=ri, column=4, value=f"{p:.4f}" if p > _P_VALUE_TINY else f"{p:.2e}"))
            else:
                sc(ws.cell(row=ri, column=3, value=""))
                sc(ws.cell(row=ri, column=4, value=""))
            ri += 1
        ri += 1

    for sev_key in ["All", "Normal", "Mild", "Moderate", "Severe"]:
        sc(ws.cell(row=ri, column=2, value=sev_names[sev_key]))
        ri += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    wb.save(TABLES_DIR / "table4.xlsx")
    logger.info("Table 4 OK")


def make_table5():
    df = pd.read_csv(TABLES_DIR / "table5_raw.csv")
    y_true = df["max_cobb"].values.astype(float)
    y_pred = df["pred_max_cobb"].values.astype(float)
    curve_type = df["curve_type"].values

    def subgroup_metrics(mask: np.ndarray, name: str) -> dict | None:
        yt, yp = y_true[mask], y_pred[mask]
        n = len(yt)
        if n == 0:
            return None
        mae = np.abs(yt-yp).mean()
        rmse = np.sqrt(((yt-yp)**2).mean())
        rng = np.random.default_rng(42)
        maeb = []
        rmseb = []
        for _ in range(2000):
            idx = rng.integers(0, n, n)
            maeb.append(np.abs(yt[idx] - yp[idx]).mean())
            rmseb.append(np.sqrt(((yt[idx] - yp[idx])**2).mean()))
        ss_res, ss_tot = ((yt-yp)**2).sum(), ((yt-yt.mean())**2).sum()
        r2 = 1 - ss_res/ss_tot if ss_tot > 0 else float('nan')
        r_val = pearsonr(yt, yp)[0] if n >= _MIN_PEARSON_SAMPLES else float('nan')
        return {
            "n": n, "gt": f"{yt.mean():.1f}", "gbdt": f"{yp.mean():.1f}",
            "mae": f"{mae:.1f} [{np.percentile(maeb,2.5):.1f},{np.percentile(maeb,97.5):.1f}]",
            "rmse": f"{rmse:.1f} [{np.percentile(rmseb,2.5):.1f},{np.percentile(rmseb,97.5):.1f}]",
            "r2": f"{r2:.2f}" if not math.isnan(r2) else "", "r": f"{r_val:.2f}" if not math.isnan(r_val) else "",
        }

    sev_mask = {label: (y_true >= lo) & (y_true < hi) for label, lo, hi in
                [("Normal", 0, 10), ("Mild", 10, 20), ("Moderate", 20, 40), ("Severe", 40, 200)]}
    ct_mask = {ct: curve_type == ct for ct in ["Thoracic", "Lumbar", "Double"]}

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 5"
    headers = ["Evaluation metrics", "GT", "GBDT", "MAE (degree)\n[95% CI]", "RMSE (degree)\n[95% CI]", "R²", "Correlation (r)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    sh(ws, 1, 1, 7)
    ws.row_dimensions[1].height = 35

    ri = 2
    m = subgroup_metrics(np.ones(len(y_true), dtype=bool), "Overall")
    if m:
        sc(ws.cell(row=ri, column=1, value="Whole cohort"), align=LEFT)
        for j, k in enumerate(["gt", "gbdt", "mae", "rmse", "r2", "r"], 2):
            sc(ws.cell(row=ri, column=j, value=m[k]))
        ri += 1

    ri += 1
    sc(ws.cell(row=ri, column=1, value="AIS severity"), font=BFONT, align=LEFT)
    for j in range(2, 8):
        sc(ws.cell(row=ri, column=j))
    ri += 1
    for name in ["Normal", "Mild", "Moderate", "Severe"]:
        m = subgroup_metrics(sev_mask[name], name)
        if m:
            sc(ws.cell(row=ri, column=1, value=f"{name} ({m['n']})"), align=LEFT)
            for j, k in enumerate(["gt", "gbdt", "mae", "rmse", "r2", "r"], 2):
                sc(ws.cell(row=ri, column=j, value=m[k]))
            ri += 1

    ri += 1
    sc(ws.cell(row=ri, column=1, value="Major curve type"), font=BFONT, align=LEFT)
    for j in range(2, 8):
        sc(ws.cell(row=ri, column=j))
    ri += 1
    for name in ["Thoracic", "Lumbar", "Double"]:
        m = subgroup_metrics(ct_mask[name], name)
        if m:
            sc(ws.cell(row=ri, column=1, value=f"{name} Curve ({m['n']})"), align=LEFT)
            for j, k in enumerate(["gt", "gbdt", "mae", "rmse", "r2", "r"], 2):
                sc(ws.cell(row=ri, column=j, value=m[k]))
            ri += 1

    ws.column_dimensions["A"].width = 30
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 20
    wb.save(TABLES_DIR / "table5.xlsx")
    logger.info("Table 5 OK")


def make_table6():
    df = pd.read_csv(TABLES_DIR / "table6_raw.csv")
    y_true = df["max_cobb"].values.astype(float)
    y_pred = df["pred_max_cobb"].values.astype(float)
    lbls = ["Normal", "Mild", "Moderate", "Severe"]
    bins = [0, 10, 20, 40, np.inf]
    tc = np.digitize(y_true, bins[1:-1])
    pc = np.digitize(y_pred, bins[1:-1])

    from sklearn.metrics import confusion_matrix
    cm = confusion_matrix(tc, pc, labels=range(4))

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 6"
    for i, h in enumerate(["Severity Grade", "Precision", "Recall", "F1-score"], 1):
        ws.cell(row=1, column=i, value=h)
    sh(ws, 1, 1, 4)

    ri = 2
    for i, lbl in enumerate(lbls):
        tp, fp = cm[i,i], cm[:,i].sum()-cm[i,i]
        fn = cm[i,:].sum()-cm[i,i]
        prec = tp/(tp+fp) if (tp+fp) else 0.0
        rec = tp/(tp+fn) if (tp+fn) else 0.0
        f1 = 2*prec*rec/(prec+rec) if (prec+rec) else 0.0
        sc(ws.cell(row=ri, column=1, value=lbl), align=LEFT)
        sc(ws.cell(row=ri, column=2, value=round(prec,3)))
        sc(ws.cell(row=ri, column=3, value=round(rec,3)))
        sc(ws.cell(row=ri, column=4, value=round(f1,3)))
        ri += 1

    acc = (tc == pc).sum() / len(y_true)
    sc(ws.cell(row=ri, column=1, value="Overall Accuracy"), font=BFONT, align=LEFT)
    sc(ws.cell(row=ri, column=2, value=round(acc,3)))
    sc(ws.cell(row=ri, column=3, value=round(acc,3)))
    sc(ws.cell(row=ri, column=4, value=round(acc,3)))
    ri += 1

    f1s, wts = [], []
    for i in range(4):
        tp, fp = cm[i,i], cm[:,i].sum()-cm[i,i]
        fn = cm[i,:].sum()-cm[i,i]
        p = tp/(tp+fp) if (tp+fp) else 0.0
        r = tp/(tp+fn) if (tp+fn) else 0.0
        f1s.append(2*p*r/(p+r) if (p+r) else 0.0)
        wts.append(cm[i,:].sum())
    w_prec = np.average([cm[i,i]/(cm[:,i].sum() or 1) for i in range(4)], weights=wts)
    w_rec = np.average([cm[i,i]/(cm[i,:].sum() or 1) for i in range(4)], weights=wts)
    w_f1 = np.average(f1s, weights=wts)

    sc(ws.cell(row=ri, column=1, value="Weighted Average"), font=BFONT, align=LEFT)
    sc(ws.cell(row=ri, column=2, value=round(w_prec,3)))
    sc(ws.cell(row=ri, column=3, value=round(w_rec,3)))
    sc(ws.cell(row=ri, column=4, value=round(w_f1,3)))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    wb.save(TABLES_DIR / "table6.xlsx")
    logger.info("Table 6 OK")


def main():
    make_table1()
    make_table2()
    make_table3()
    make_table4()
    make_table5()
    make_table6()
    logger.info("\n全部 Excel 表格已生成。")


if __name__ == "__main__":
    make_table1()
    make_table2()
    make_table3()
    make_table4()
    make_table5()
    make_table6()
    logger.info("\n全部 Excel 表格已生成。")
